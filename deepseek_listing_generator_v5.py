# -*- coding: utf-8 -*-
"""
汽配零件 Listing 自动生成工具 (浏览器版 v5 - 最终修正版)
通过 Selenium 控制 Edge 浏览器操作 DeepSeek 网

页版，生成亚马逊 listing 并回写 Excel。

修复内容（v5）：
1. 修正标题提取逻辑，提取完整标题
2. 修正段落映射，确保各字段对应正确列
3. 保留原始 OE 号，不过滤关键数字
"""
import openpyxl
import shutil
import time
import re
import os

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# ============================================================
# 配置区 - 根据需要修改
# ============================================================
EXCEL_PATH = r"C:\Users\Administrator\Desktop\test2.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\test2_output.xlsx"
DATA_START_ROW = 7
WAIT_TIMEOUT = 120
REPLY_STABLE_COUNT = 4
CHECK_INTERVAL = 3
ACTION_DELAY = 2
MAX_RETRIES = 3
DEBUG_MODE = True

# 列映射
COL_MAP = {
    'title': 7,              # G
    'description': 30,       # AD
    'bullet1': 31,           # AE
    'bullet2': 32,           # AF
    'bullet3': 33,           # AG
    'bullet4': 34,           # AH
    'bullet5': 35,           # AI
    'keyword1': 36,          # AJ
    'keyword2': 37,          # AK
    'keyword3': 38,          # AL
    'keyword4': 39,          # AM
    'keyword5': 40,          # AN
}


# ============================================================
# 修正后的解析器
# ============================================================
def parse_deepseek_response(text, oe_number):
    """根据 DeepSeek 写作思路说明解析回复"""
    if not text or len(text) < 100:
        return None

    result = {
        'oe': oe_number,
        'title': '',
        'description': '',
        'bullet1': '',
        'bullet2': '',
        'bullet3': '',
        'bullet4': '',
        'bullet5': '',
        'keyword1': '',
        'keyword2': '',
        'keyword3': '',
        'keyword4': '',
        'keyword5': '',
    }

    lines = text.split('\n')

    # 定位表格区域
    header_idx = -1
    data_idx = -1

    for i, line in enumerate(lines):
        stripped = line.strip()
        if re.match(r'^A\s+G\s+AD\s+AE\s+AF\s+AG\s+AH\s+AI\s+AJ\s+AK\s+AL\s+AM\s+AN$', stripped):
            header_idx = i
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line.startswith(oe_number):
                    data_idx = i + 1
            break

    if data_idx == -1:
        print(f"    ✗ 未找到表格数据行")
        return None

    # 提取标题（G列）
    data_line = lines[data_idx].strip()
    fields = data_line.split()

    if len(fields) < 2:
        print(f"    ✗ 数据字段不足")
        return None

    # 从字段 1 开始提取到 "Vehicle Fitment:" 之前
    title_fields = fields[1:]

    title_end_idx = -1
    for i, field in enumerate(title_fields):
        if field == 'Vehicle' and i + 1 < len(title_fields) and title_fields[i+1] == 'Fitment:':
            title_end_idx = i
            break

    if title_end_idx > 0:
        title_fields = title_fields[:title_end_idx]

    result['title'] = ' '.join(title_fields)

    # 提取段落内容
    # AD 列: Vehicle Fitment: -> 描述
    vehicle_section = extract_section(text, 'Vehicle Fitment:',
                                      ['Direct Fitment', 'OEM Cross Reference', 'Precise', 'Plug', 'Durable', '1-Year', '写作'])
    if vehicle_section:
        result['description'] = clean_text(vehicle_section)

    # AE 列: Direct Fitment: -> 卖点1
    direct_section = extract_section(text, 'Direct Fitment:',
                                    ['OEM Cross Reference', 'Precise', 'Plug', 'Durable', '1-Year', '写作'])
    if direct_section:
        result['bullet1'] = clean_text(direct_section)

    # AF 列: OEM Cross Reference: -> 卖点2
    oem_section = extract_section(text, 'OEM Cross Reference:',
                                  ['Precise', 'Plug', 'Durable', '1-Year', '写作'])
    if oem_section:
        result['bullet2'] = clean_text(oem_section)

    # AG 列: Precise ...: -> 卖点3
    precise_patterns = ['Precise Signal Output:', 'Precise Air Volume Measurement:', 'Precise:']
    precise_section = extract_section(text, precise_patterns,
                                      ['Plug', 'Durable', '1-Year', '写作'])
    if precise_section:
        result['bullet3'] = clean_text(precise_section)

    # AH 列: Plug & Play Installation: -> 卖点4
    plug_section = extract_section(text, ['Plug & Play Installation:', 'Plug Play Installation:'],
                                   ['Durable', '1-Year', '写作'])
    if plug_section:
        result['bullet4'] = clean_text(plug_section)

    # AI 列: Durable Construction: + 1-Year Warranty: -> 卖点5
    durable_section = extract_section(text, 'Durable Construction:',
                                      ['1-Year', '写作'])
    if durable_section:
        durable_text = clean_text(durable_section)
        result['bullet5'] = durable_text

    warranty_section = extract_section(text, '1-Year Warranty:',
                                      ['写作', '关键词'])
    if warranty_section:
        warranty_text = clean_text(warranty_section)
        if result['bullet5']:
            result['bullet5'] = result['bullet5'] + ' ' + warranty_text
        else:
            result['bullet5'] = warranty_text

    # 关键词
    keywords_pattern = rf'{re.escape(oe_number)}\s+([\w\s\-\(\)]+?)\s*$'
    keywords_match = re.search(keywords_pattern, text, re.MULTILINE)
    if keywords_match:
        keywords_text = keywords_match.group(1).strip()
        keyword_list = [k.strip() for k in keywords_text.split() if k.strip()]
        for i, kw in enumerate(keyword_list[:5]):
            result[f'keyword{i+1}'] = kw

    # 验证
    if result['title'] and (result['description'] or result['bullet1']):
        print(f"    ✓ 解析成功!")
        return result
    else:
        print(f"    ✗ 解析失败: 缺少关键字段")
        return None


def extract_section(text, keywords, stop_keywords):
    """从文本中提取段落"""
    if isinstance(keywords, str):
        keywords = [keywords]

    escaped_keywords = [re.escape(k) for k in keywords]
    pattern = '|'.join(escaped_keywords)

    match = re.search(pattern, text, re.IGNORECASE)
    if not match:
        return None

    start_pos = match.end()
    stop_pos = len(text)

    for stop_kw in stop_keywords:
        stop_match = re.search(rf'\n\s*{re.escape(stop_kw)}', text[start_pos:], re.IGNORECASE)
        if stop_match:
            stop_pos = start_pos + stop_match.start()
            break

    content = text[start_pos:stop_pos].strip()
    # 移除引用标记（但保留 OE 号等）
    content = re.sub(r'\n\s*\d+\s*\n', '\n', content)
    content = re.sub(r'\s+', ' ', content)
    return content


def clean_text(text):
    """清理文本，但保留 OE 号等关键信息"""
    if not text:
        return ''
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


# ============================================================
# Excel 操作
# ============================================================
def get_oe_numbers(filepath):
    """从 A 列读取 OE 号"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    result = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            val = str(val).strip()
            if val:
                val = re.sub(r'[-\s]+[Pp]\d+$', '', val)
                result.append((row, val))
    return result


def write_results(filepath, output_path, results):
    """将解析结果写入 Excel"""
    shutil.copy2(filepath, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for row, data in results.items():
        if not data:
            continue
        for field, col in COL_MAP.items():
            val = data.get(field, '')
            if val:
                ws.cell(row=row, column=col, value=val)

    wb.save(output_path)


# ============================================================
# 浏览器操作
# ============================================================
class DeepSeekBot:
    def __init__(self):
        self.driver = None

    def start(self):
        """启动 Edge"""
        print("启动 Edge 浏览器...")
        opts = Options()
        opts.add_argument('--start-maximized')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('detach', True)
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])

        self.driver = webdriver.Edge(options=opts)
        self.driver.get("https://chat.deepseek.com")
        print(f"已打开 DeepSeek，等待页面加载...")
        time.sleep(8)

    def _find_input(self):
        """查找输入框"""
        selectors = [
            'textarea.dsb-textarea',
            'textarea[placeholder*="输入"]',
            'textarea[placeholder*="message"]',
            'div[contenteditable="true"]',
            'textarea',
            '#chat-input textarea',
            'div[class*="input"] textarea',
        ]
        for sel in selectors:
            try:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        return el
            except Exception:
                continue
        return None

    def _find_send_button(self):
        """查找发送按钮"""
        selectors = [
            'button[data-testid="send-button"]',
            'button[type="submit"]',
            'button[class*="send"]',
            'div[class*="send"] button',
            'button svg[class*="send"]',
            'button[aria-label*="send"]',
            'button[aria-label*="发送"]',
        ]
        for sel in selectors:
            try:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        return el
            except Exception:
                continue
        return None

    def _click_new_chat(self):
        """尝试开新对话"""
        selectors = [
            'button[class*="new"]',
            'a[class*="new"]',
            '[class*="new-chat"]',
            'div[class*="sidebar"] button:first-child',
            'button[aria-label*="new"]',
        ]
        for sel in selectors:
            try:
                btn = self.driver.find_element(By.CSS_SELECTOR, sel)
                if btn.is_displayed():
                    btn.click()
                    time.sleep(2)
                    return True
            except Exception:
                continue
        return False

    def _get_last_message_text(self):
        """获取最后一条消息的文本内容"""
        try:
            messages = self.driver.find_elements(By.CSS_SELECTOR,
                                                'div[class*="message"], div[class*="chat-bubble"], div[class*="conversation"]')

            if messages:
                last_msg = messages[-1]
                return last_msg.text

            body_text = self.driver.find_element(By.TAG_NAME, 'body').text
            oe_pattern = r'[A-Z0-9\-\.]+(?=[^\n]*?compatible|已收到您的OE号)'
            matches = list(re.finditer(oe_pattern, body_text))

            if matches:
                last_pos = matches[-1].end()
                return body_text[last_pos:]

            return body_text

        except Exception as e:
            print(f"    获取消息失败: {str(e)}")
            return ""

    def send_and_get_reply(self, oe_number, max_retries=MAX_RETRIES):
        """发送 OE 号并获取完整回复"""
        retries = 0
        while retries <= max_retries:
            try:
                if not self._click_new_chat():
                    print("    警告: 无法开新对话")

                input_el = self._find_input()
                if not input_el:
                    print(f"    找不到输入框 (尝试 {retries+1}/{max_retries})!")
                    retries += 1
                    time.sleep(2)
                    continue

                input_el.clear()
                input_el.click()
                time.sleep(0.5)
                input_el.send_keys(oe_number)
                time.sleep(ACTION_DELAY)

                send_btn = self._find_send_button()
                if send_btn:
                    send_btn.click()
                else:
                    input_el.send_keys(Keys.RETURN)

                return self._wait_for_reply()

            except Exception as e:
                print(f"    发送失败 (尝试 {retries+1}/{max_retries}): {str(e)}")
                retries += 1
                time.sleep(2)
                if retries > max_retries:
                    print("    超过最大重试次数，跳过此条目")
                    return None

    def _wait_for_reply(self):
        """等待 DeepSeek 回复完成"""
        print(f"    等待回复 (最长 {WAIT_TIMEOUT}s)...")
        start = time.time()
        last_text = ""
        last_length = 0
        stable_count = 0

        while time.time() - start < WAIT_TIMEOUT:
            time.sleep(CHECK_INTERVAL)
            current_text = self._get_last_message_text()
            current_length = len(current_text)

            if 'compatible' in current_text.lower() or 'fitment' in current_text.lower() or len(current_text) > 200:
                if current_length == last_length and current_length > 0:
                    stable_count += 1
                    if stable_count >= REPLY_STABLE_COUNT:
                        elapsed = int(time.time() - start)
                        print(f"    回复完成 ({elapsed}s, {current_length} 字符)")
                        return current_text
                else:
                    stable_count = 0
                    last_length = current_length

                if current_length > 0:
                    dots = "." * (int(time.time() - start) % 4)
                    print(f"    生成中{dots} ({int(time.time() - start)}s, {current_length} 字符)")

            last_text = current_text

        print(f"    超时，使用已收到的内容")
        return last_text

    def close(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("  汽配 Listing 自动生成工具 v5 (最终修正版)")
    print("  DeepSeek 浏览器自动化版")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"\nExcel 文件不存在: {EXCEL_PATH}")
        return

    # 读取 OE 号
    print(f"\n读取 Excel: {EXCEL_PATH}")
    oe_list = get_oe_numbers(EXCEL_PATH)
    print(f"找到 {len(oe_list)} 个 OE 号")
    for i, (row, oe) in enumerate(oe_list[:5]):
        print(f"  {i+1}. [Row {row}] {oe}")
    if len(oe_list) > 5:
        print(f"  ... 还有 {len(oe_list) - 5} 个")

    # 选择范围
    print(f"\n处理范围:")
    print(f"  1. 全部 ({len(oe_list)} 个)")
    print(f"  2. 测试前 3 个")
    print(f"  3. 自定义 (如 5-10)")
    choice = input("选择 (1/2/3): ").strip()

    if choice == '2':
        work = oe_list[:3]
    elif choice == '3':
        try:
            rng = input("范围 (如 5-10): ").strip().split('-')
            work = oe_list[int(rng[0])-1:int(rng[1])]
        except Exception:
            print("格式错误，处理全部")
            work = oe_list
    else:
        work = oe_list

    print(f"\n将处理 {len(work)} 个")

    # 启动浏览器
    bot = DeepSeekBot()
    try:
        bot.start()

        input("\n>>> 浏览器已打开，请确认 DeepSeek 已登录且 Prompt 已配置好。按 Enter 继续...")

        results = {}
        ok = 0
        fail = 0

        for i, (row, oe) in enumerate(work):
            print(f"\n{'='*60}")
            print(f"[{i+1}/{len(work)}] OE: {oe} (Row {row})")
            print(f"{'='*60}")

            reply = bot.send_and_get_reply(oe)
            if not reply:
                print("    ✗ 未收到回复")
                fail += 1
                continue

            # 调试: 保存原始回复
            if DEBUG_MODE:
                debug_file = f"C:\\Users\\Administrator\\Desktop\\debug_v5_row{row}.txt"
                with open(debug_file, 'w', encoding='utf-8') as f:
                    f.write(f"OE Number: {oe}\n")
                    f.write(f"Row: {row}\n")
                    f.write("=" * 80 + "\n")
                    f.write(reply)
                print(f"    调试文件已保存: {debug_file}")

            # 解析
            data = parse_deepseek_response(reply, oe)
            if data and data.get('title'):
                results[row] = data
                ok += 1
                print(f"    ✓ 成功!")
                print(f"      标题: {data['title'][:60]}...")
                print(f"      描述: {data['description'][:60]}...")
                print(f"      卖点1: {data['bullet1'][:60]}...")
            else:
                fail += 1
                print(f"    ✗ 解析失败!")

            if i < len(work) - 1:
                time.sleep(ACTION_DELAY)

        # 写入结果
        if results:
            write_results(EXCEL_PATH, OUTPUT_PATH, results)
            print(f"\n{'='*60}")
            print(f"✓ 结果已保存: {OUTPUT_PATH}")
        else:
            print(f"\n{'='*60}")
            print(f"✗ 没有成功解析的数据")

        # 总结
        print(f"\n{'='*60}")
        print(f"  完成! 成功: {ok} | 失败: {fail} | 总计: {len(work)}")
        print(f"{'='*60}")

    except KeyboardInterrupt:
        print("\n用户中断")
    except Exception as e:
        print(f"\n出错: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("\n浏览器保持打开，可手动检查。")


if __name__ == "__main__":
    main()
