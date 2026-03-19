# -*- coding: utf-8 -*-
"""
汽配 Listing 自动生成工具 v4.1（table td 提取版 - 修复）

核心逻辑：
1. 从 DeepSeek <table> <td> 直接提取数据
2. DeepSeek 的 table 可能有 14 列（AI 和 warranty 分开了）
3. 合并 td[7]+td[8] 到 AI 列，td[9]-td[13] 映射到 AJ-AN
4. 如果只有 13 列，按标准映射
"""
import openpyxl
import shutil
import time
import re
import os

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# ============================================================
# 配置区
# ============================================================
EXCEL_PATH = r"C:\Users\Administrator\Desktop\test2.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\test2_output.xlsx"
DATA_START_ROW = 7
WAIT_TIMEOUT = 120
REPLY_STABLE_COUNT = 4
CHECK_INTERVAL = 3
ACTION_DELAY = 2
MAX_RETRIES = 3
DEBUG_MODE = True

# 列映射：列名 → openpyxl 列号
COL_MAP = {
    'A': 1, 'G': 7, 'AD': 30,
    'AE': 31, 'AF': 32, 'AG': 33, 'AH': 34, 'AI': 35,
    'AJ': 36, 'AK': 37, 'AL': 38, 'AM': 39, 'AN': 40,
}

# 期望的表头列名（按顺序，13 个）
EXPECTED_HEADERS = ['A', 'G', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']


# ============================================================
# 从 <table> <td> 提取数据
# ============================================================
def extract_from_table(driver, oe_number):
    """
    从 DeepSeek 页面最后一个 table 的 td 元素提取数据。
    DeepSeek 的 table 结构（已确认）：
      <thead><tr><th>A</th><th>G</th>...<th>AN</th></tr></thead>
      <tbody><tr><td>OE号</td><td>标题</td><td>描述</td>...<td>关键词</td></tr></tbody>
    
    注意：DeepSeek 可能把 AI 列的 warranty 部分拆成独立 td（共14列），
    也可能只有 13 列。代码兼容两种情况。
    """
    try:
        tables = driver.find_elements(By.TAG_NAME, 'table')
        if not tables:
            print("    未找到 <table>")
            return None

        # 从后往前找匹配的 table
        for table in reversed(tables):
            thead = table.find_elements(By.TAG_NAME, 'thead')
            tbody = table.find_elements(By.TAG_NAME, 'tbody')
            if not tbody:
                continue

            # 验证表头
            ths = []
            if thead:
                ths = thead[0].find_elements(By.TAG_NAME, 'th')
            if not ths:
                # 也可能没有 thead，直接看第一行 tr
                first_tr = table.find_elements(By.TAG_NAME, 'tr')
                if first_tr:
                    ths = first_tr[0].find_elements(By.TAG_NAME, 'th')
                    if not ths:
                        ths = first_tr[0].find_elements(By.TAG_NAME, 'td')
            
            header_texts = [t.text.strip().upper() for t in ths]
            
            # 匹配条件：包含 A, G, AD
            if 'A' not in header_texts or 'G' not in header_texts or 'AD' not in header_texts:
                continue

            if len(header_texts) < 10:
                continue

            # 找数据行
            data_trs = tbody[0].find_elements(By.TAG_NAME, 'tr')
            if not data_trs:
                continue

            target_tr = data_trs[0]
            tds = target_tr.find_elements(By.TAG_NAME, 'td')

            if not tds or len(tds) < 10:
                print(f"    table 只有 {len(tds)} 个 td，跳过")
                continue

            print(f"    找到 table: {len(ths)} 表头, {len(tds)} 数据列")

            # 提取所有 td 文本
            td_texts = []
            for td in tds:
                text = td.text.strip()
                text = re.sub(r'\s+', ' ', text)  # 合并换行
                td_texts.append(text)

            # 构建结果
            result = {
                'A': '', 'G': '', 'AD': '', 'AE': '', 'AF': '', 'AG': '',
                'AH': '', 'AI': '', 'AJ': '', 'AK': '', 'AL': '', 'AM': '', 'AN': '',
            }

            num_cols = len(td_texts)

            if num_cols == 14:
                # 14 列：AI 和 warranty 分开了
                # td[0]=A, td[1]=G, td[2]=AD, td[3]=AE, td[4]=AF, td[5]=AG,
                # td[6]=AH, td[7]=AI(主体), td[8]=warranty(合并到AI),
                # td[9]-td[13] = AJ-AN
                mapping_14 = {
                    0: 'A', 1: 'G', 2: 'AD', 3: 'AE', 4: 'AF', 5: 'AG',
                    6: 'AH',
                    7: 'AI',  # AI 主体
                    9: 'AJ', 10: 'AK', 11: 'AL', 12: 'AM', 13: 'AN',
                }
                for idx, col_name in mapping_14.items():
                    if idx < num_cols:
                        result[col_name] = td_texts[idx]
                # td[8] = warranty，合并到 AI
                if 8 < num_cols:
                    ai = result.get('AI', '')
                    warranty = td_texts[8]
                    if ai and warranty:
                        result['AI'] = ai + ' ' + warranty
                    elif warranty:
                        result['AI'] = warranty

            elif num_cols == 13:
                # 标准的 13 列
                for i, col_name in enumerate(EXPECTED_HEADERS):
                    if i < num_cols:
                        result[col_name] = td_texts[i]

            elif num_cols >= 10:
                # 其他情况：尽量映射
                for i, col_name in enumerate(EXPECTED_HEADERS):
                    if i < num_cols:
                        result[col_name] = td_texts[i]
            else:
                return None

            # 验证
            if result['G'] and len(result['G']) > 10:
                filled = sum(1 for v in result.values() if v)
                print(f"    table 提取成功! {filled}/13 字段, G: {len(result['G'])}c")
                return result
            else:
                print(f"    table G 列太短 ({len(result['G'])}c)")
                return None

        print("    所有 table 都不匹配")
        return None

    except Exception as e:
        print(f"    table 提取异常: {str(e)}")
        return None


# ============================================================
# 文本解析（fallback，保留但通常不需要）
# ============================================================
HEADER_RE = re.compile(
    r'^A\s+G\s+AD\s+AE\s+AF\s+AG\s+AH\s+AI\s+AJ\s+AK\s+AL\s+AM\s+AN\s*$',
    re.IGNORECASE
)

def parse_deepseek_text(text, oe_number):
    if not text or len(text) < 100:
        return None
    lines = text.split('\n')
    cleaned = []
    started = False
    for line in lines:
        s = line.strip()
        if HEADER_RE.match(s):
            started = True; continue
        if not started: continue
        if re.match(r'^(已读取|Read).*网页', s, re.I): continue
        if re.match(r'^\d+$', s): continue
        if s: cleaned.append(s)
    merged = ' '.join(cleaned)
    merged = re.sub(r'\)\s+[\d\s]+([,;.])', r')\1', merged)
    merged = re.sub(r'\s+', ' ', merged).strip()
    
    result = {
        'A': oe_number, 'G': '', 'AD': '', 'AE': '', 'AF': '', 'AG': '',
        'AH': '', 'AI': '', 'AJ': '', 'AK': '', 'AL': '', 'AM': '', 'AN': '',
    }
    pos = merged.find(oe_number)
    if pos < 0: return None
    vf = re.search(r'Vehicle\s+Fitment:', merged, re.I)
    if not vf: return None
    result['G'] = merged[pos+len(oe_number):vf.start()].strip()
    result['G'] = re.sub(r'^\d{4,5}[-]\d{4,5}\s*', '', result['G']).strip()
    result['G'] = re.sub(r'\s+\d{4,5}[-]\d{4,5}\s*$', '', result['G']).strip()
    if not result['G']: return None
    
    oes = list(re.finditer(re.escape(oe_number), merged))
    last_oe_end = oes[-1].end() if oes else len(merged)
    vf_end = vf.end()
    
    bps = [r'Verify\s+vehicle', r'Direct\s+fit\s+for', r'Replaces\s+OEM',
           r'Provides\s+accurate', r'Measures\s+intake', r'Plug\s*&?\s*play',
           r'Delivers\s+accurate', r'Durable', r'Corrosion-resistant',
           r'100%\s+tested', r'1-?year\s+warranty', r'Meets\s+OE', r'Meets\s+Toyota',
           r'No\s+splicing', r'Precise\s+[SA]', r'Factory-Grade']
    after = merged[vf_end:]
    found = []
    for p in bps:
        for m in re.finditer(p, after, re.I):
            found.append(vf_end + m.start())
    found = sorted(set(found))
    ff = []
    for f in found:
        if not ff or f - ff[-1] > 5:
            ff.append(f)
    ff = [f for f in ff if vf_end < f < last_oe_end]
    
    if ff:
        result['AD'] = merged[vf_end:ff[0]].strip()
    else:
        result['AD'] = merged[vf_end:last_oe_end].strip()
    result['AD'] = re.sub(r'\s*\.?\s*$', '', result['AD'])
    
    cols = ['AE','AF','AG','AH','AI']
    for i in range(min(5, len(ff))):
        s = ff[i]
        e = ff[i+1] if i+1 < len(ff) else last_oe_end
        if i < len(cols):
            result[cols[i]] = merged[s:e].strip()
    
    if last_oe_end < len(merged):
        kw = merged[last_oe_end:].strip()
        kw = re.sub(r'(已读取|Read).*$', '', kw, flags=re.DOTALL).strip()
        kws = [k.strip() for k in kw.split() if len(k.strip()) > 1]
        for i, k in enumerate(kws[:5]):
            result[['AJ','AK','AL','AM','AN'][i]] = k
    
    return result if result['G'] else None


# ============================================================
# Excel 操作
# ============================================================
def get_oe_numbers(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    result = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            val = str(val).strip()
            if val:
                val = re.sub(r'[-\s]+[Pp]\d+$', '', val)
                result.append((row, val))
    return result


def write_results(filepath, output_path, results):
    shutil.copy2(filepath, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    print(f"\n写入 Excel：")
    for row, data in results.items():
        print(f"  Row {row}:")
        for col_name, col_num in COL_MAP.items():
            val = data.get(col_name, '')
            if val:
                ws.cell(row=row, column=col_num, value=val)
                print(f"    {col_name:4} -> 列{col_num} ({len(val):3}c)")
    wb.save(output_path)
    print(f"\n>>> 已保存: {output_path}")


# ============================================================
# 浏览器操作
# ============================================================
class DeepSeekBot:
    def __init__(self):
        self.driver = None

    def start(self):
        print("启动 Edge...")
        opts = Options()
        opts.add_argument('--start-maximized')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('detach', True)
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.driver = webdriver.Edge(options=opts)
        self.driver.get("https://chat.deepseek.com")
        print("等待加载...")
        time.sleep(8)

    def _find_input(self):
        for sel in ['textarea.dsb-textarea', 'textarea[placeholder*="输入"]',
                     'textarea[placeholder*="message"]', 'div[contenteditable="true"]',
                     'textarea', 'div[class*="input"] textarea']:
            try:
                for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.is_enabled():
                        return el
            except: continue
        return None

    def _find_send_button(self):
        for sel in ['button[data-testid="send-button"]', 'button[type="submit"]',
                     'button[class*="send"]', 'div[class*="send"] button',
                     'button[aria-label*="send"]', 'button[aria-label*="发送"]']:
            try:
                for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.is_enabled():
                        return el
            except: continue
        return None

    def _click_new_chat(self):
        for sel in ['button[class*="new"]', 'a[class*="new"]', '[class*="new-chat"]',
                     'div[class*="sidebar"] button:first-child', 'button[aria-label*="new"]']:
            try:
                btn = self.driver.find_element(By.CSS_SELECTOR, sel)
                if btn.is_displayed():
                    btn.click()
                    time.sleep(2)
                    return True
            except: continue
        return False

    def _wait_for_reply(self):
        print(f"    等待回复 (最长 {WAIT_TIMEOUT}s)...")
        start = time.time()
        last_len = 0
        stable = 0

        while time.time() - start < WAIT_TIMEOUT:
            time.sleep(CHECK_INTERVAL)
            try:
                tables = self.driver.find_elements(By.TAG_NAME, 'table')
                # 找到最后一个有 thead 的 table 的 td 数量
                td_count = 0
                for t in reversed(tables):
                    ths = t.find_elements(By.TAG_NAME, 'th')
                    header_texts = [th.text.strip().upper() for th in ths]
                    if 'A' in header_texts and 'G' in header_texts and 'AD' in header_texts:
                        tbody = t.find_elements(By.TAG_NAME, 'tbody')
                        if tbody:
                            trs = tbody[0].find_elements(By.TAG_NAME, 'tr')
                            if trs:
                                td_count = len(trs[0].find_elements(By.TAG_NAME, 'td'))
                        break

                body_len = len(self.driver.find_element(By.TAG_NAME, 'body').text)
            except:
                td_count = 0
                body_len = 0

            # 有 table 且数据稳定 = 回复完成
            current_marker = td_count * 10000 + body_len % 10000
            if td_count > 0 and current_marker == last_len:
                stable += 1
                if stable >= REPLY_STABLE_COUNT:
                    elapsed = int(time.time() - start)
                    print(f"    回复完成 ({elapsed}s, {td_count} td)")
                    return True
            else:
                stable = 0
            last_len = current_marker

            dots = "." * (int(time.time() - start) % 4)
            print(f"    生成中{dots} ({int(time.time() - start)}s, td={td_count})")

        print(f"    超时")
        return False

    def send_and_get_reply(self, oe_number):
        retries = 0
        while retries <= MAX_RETRIES:
            try:
                if not self._click_new_chat():
                    print("    ! 无法开新对话，尝试继续")

                input_el = self._find_input()
                if not input_el:
                    print(f"    找不到输入框 ({retries+1}/{MAX_RETRIES})")
                    retries += 1; time.sleep(2); continue

                input_el.clear(); input_el.click(); time.sleep(0.5)
                input_el.send_keys(oe_number); time.sleep(ACTION_DELAY)

                send_btn = self._find_send_button()
                if send_btn:
                    send_btn.click()
                else:
                    input_el.send_keys(Keys.RETURN)

                if self._wait_for_reply():
                    return True
                else:
                    retries += 1
                    print(f"    超时，重试...")
                    time.sleep(2)

            except Exception as e:
                print(f"    失败 ({retries+1}/{MAX_RETRIES}): {str(e)}")
                retries += 1; time.sleep(2)

        print("    超过最大重试次数")
        return False

    def close(self):
        if self.driver:
            try: self.driver.quit()
            except: pass


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("  汽配 Listing 自动生成工具 v4.1")
    print("  从 DeepSeek <table> <td> 直接提取数据")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"\nExcel 不存在: {EXCEL_PATH}"); return

    print(f"\n读取 Excel: {EXCEL_PATH}")
    oe_list = get_oe_numbers(EXCEL_PATH)
    print(f"找到 {len(oe_list)} 个 OE 号")

    print(f"\n处理范围:")
    print(f"  1. 全部 ({len(oe_list)} 个)")
    print(f"  2. 测试前 3 个")
    choice = input("选择 (1/2): ").strip()
    work = oe_list[:3] if choice == '2' else oe_list
    print(f"\n将处理 {len(work)} 个")

    bot = DeepSeekBot()
    try:
        bot.start()
        input("\n>>> 浏览器已打开，确认 DeepSeek 已登录且 Prompt 已配置好。按 Enter 继续...")

        results = {}
        ok = 0
        fail = 0
        method_used = {'table': 0, 'text': 0}

        for i, (row, oe) in enumerate(work):
            print(f"\n{'='*60}")
            print(f"[{i+1}/{len(work)}] OE: {oe} (Row {row})")
            print(f"{'='*60}")

            if not bot.send_and_get_reply(oe):
                print("    X 未收到回复"); fail += 1; continue

            # 调试
            if DEBUG_MODE:
                try:
                    tables = bot.driver.find_elements(By.TAG_NAME, 'table')
                    html = tables[-1].get_attribute('outerHTML') if tables else 'NO TABLE'
                    debug_file = os.path.join(
                        r"C:\Users\Administrator\Desktop\deepseek_listing_tool",
                        f"debug_v41_row{row}.html"
                    )
                    with open(debug_file, 'w', encoding='utf-8') as f:
                        f.write(f"OE: {oe}\nRow: {row}\nTables on page: {len(tables)}\n{'='*80}\n")
                        f.write(html)
                except: pass

            # 方法1：table td 提取
            data = extract_from_table(bot.driver, oe)
            if data:
                method_used['table'] += 1
            else:
                # 方法2：文本 fallback
                print("    → fallback 文本解析")
                try:
                    body_text = bot.driver.find_element(By.TAG_NAME, 'body').text
                    data = parse_deepseek_text(body_text, oe)
                    if data:
                        method_used['text'] += 1
                except Exception as e:
                    print(f"    fallback 也失败: {str(e)}")

            if data:
                results[row] = data; ok += 1
                print(f"    OK!")
                for col in ['G', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']:
                    val = data.get(col, '')
                    if val:
                        print(f"      {col}: {len(val):4}c | {val[:70]}{'...' if len(val)>70 else ''}")
            else:
                fail += 1
                print(f"    X 失败!")

            if i < len(work) - 1:
                time.sleep(ACTION_DELAY)

        if results:
            write_results(EXCEL_PATH, OUTPUT_PATH, results)

        print(f"\n{'='*60}")
        print(f"  完成! 成功: {ok} | 失败: {fail} | 总计: {len(work)}")
        print(f"  提取方式: table={method_used['table']}, text={method_used['text']}")
        print(f"{'='*60}")

    except KeyboardInterrupt:
        print("\n用户中断")
    except Exception as e:
        print(f"\n出错: {e}")
        import traceback; traceback.print_exc()
    finally:
        print("\n浏览器保持打开。")


if __name__ == "__main__":
    main()