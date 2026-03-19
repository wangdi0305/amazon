# -*- coding: utf-8 -*-
"""
离线测试脚本 - 使用现有的调试文件测试解析逻辑
无需浏览器，直接解析已有的 DeepSeek 响应
"""
import openpyxl
import shutil
import time
import re
import os
import sys

# ============================================================
# 配置
# ============================================================
EXCEL_PATH = r"C:\Users\Administrator\Desktop\test2.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\test2_output_offline.xlsx"
DEBUG_DIR = r"C:\Users\Administrator\Desktop"

# 列映射
COL_MAP = {
    'title': 7,              # G
    'description': 30,       # AD
    'bullet1': 31,           # AE
    'bullet2': 32,           # AF
    'bullet3': 33,           # AG
    'bullet4': 34,           # AH
    'bullet5': 35,           # AI
    'keyword1': 36,          # AJ
    'keyword2': 37,          # AK
    'keyword3': 38,          # AL
    'keyword4': 39,          # AM
    'keyword5': 40,          # AN
}


# ============================================================
# 解析器（从 v4 复制）
# ============================================================
def parse_deepseek_response(text, oe_number):
    """智能解析 DeepSeek 的回复"""
    if not text or len(text) < 100:
        return None

    result = {
        'oe': oe_number,
        'title': '',
        'description': '',
        'bullet1': '',
        'bullet2': '',
        'bullet3': '',
        'bullet4': '',
        'bullet5': '',
        'keyword1': '',
        'keyword2': '',
        'keyword3': '',
        'keyword4': '',
        'keyword5': '',
    }

    lines = text.split('\n')

    # 定位表格区域
    header_idx = -1
    data_idx = -1

    for i, line in enumerate(lines):
        stripped = line.strip()
        if re.match(r'^A\s+G\s+AD\s+AE\s+AF\s+AG\s+AH\s+AI\s+AJ\s+AK\s+AL\s+AM\s+AN$', stripped):
            header_idx = i
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line.startswith(oe_number):
                    data_idx = i + 1
            break

    if data_idx == -1:
        return None

    # 提取表格数据
    data_line = lines[data_idx].strip()
    fields = data_line.split()

    if len(fields) < 3:
        return None

    # 提取标题
    title_end_idx = -1
    for i, field in enumerate(fields):
        if field.lower() in ['compatible', 'for', 'fits']:
            title_end_idx = i
            break

    if title_end_idx == -1 or title_end_idx < 2:
        return None

    title_fields = fields[2:title_end_idx + 1]
    result['title'] = ' '.join(title_fields)

    # 提取段落内容
    vehicle_section = extract_section(text, 'Vehicle Fitment:',
                                      ['Direct Fitment', 'OEM Cross Reference', 'Precise', 'Plug', 'Durable', '1-Year', '写作'])
    if vehicle_section:
        result['description'] = clean_text(vehicle_section)

    oem_section = extract_section(text, 'OEM Cross Reference:',
                                  ['Precise', 'Plug', 'Durable', '1-Year', '写作'])
    if oem_section:
        result['bullet2'] = clean_text(oem_section)

    precise_section = extract_section(text, 'Precise Signal Output:',
                                      ['Plug', 'Durable', '1-Year', '写作'])
    if precise_section:
        result['bullet3'] = clean_text(precise_section)

    plug_section = extract_section(text, ['Plug & Play Installation:', 'Plug Play Installation:'],
                                   ['Durable', '1-Year', '写作'])
    if plug_section:
        result['bullet4'] = clean_text(plug_section)

    durable_section = extract_section(text, 'Durable Construction:',
                                      ['1-Year', '写作'])
    if durable_section:
        durable_text = clean_text(durable_section)
        if result['bullet5']:
            result['bullet5'] = result['bullet5'] + ' ' + durable_text
        else:
            result['bullet5'] = durable_text

    warranty_section = extract_section(text, '1-Year Warranty:',
                                      ['写作', '关键词'])
    if warranty_section:
        warranty_text = clean_text(warranty_section)
        if result['bullet5']:
            result['bullet5'] = result['bullet5'] + ' ' + warranty_text
        else:
            result['bullet5'] = warranty_text

    # 提取适配车型
    if result['description']:
        sentences = result['description'].split('.')
        if sentences:
            first_sentence = sentences[0].strip()
            if 'compatible' in first_sentence.lower():
                result['bullet1'] = first_sentence
                result['description'] = '. '.join(sentences[1:]).strip()

    # 提取关键词
    keywords_pattern = rf'{re.escape(oe_number)}\s+([\w\s\-\(\)]+?)\s*$'
    keywords_match = re.search(keywords_pattern, text, re.MULTILINE)
    if keywords_match:
        keywords_text = keywords_match.group(1).strip()
        keyword_list = [k.strip() for k in keywords_text.split() if k.strip()]
        for i, kw in enumerate(keyword_list[:5]):
            result[f'keyword{i+1}'] = kw

    if result['title'] and (result['description'] or result['bullet1']):
        return result
    else:
        return None


def extract_section(text, keywords, stop_keywords):
    """从文本中提取段落"""
    if isinstance(keywords, str):
        keywords = [keywords]

    escaped_keywords = [re.escape(k) for k in keywords]
    pattern = '|'.join(escaped_keywords)

    match = re.search(pattern, text, re.IGNORECASE)
    if not match:
        return None

    start_pos = match.end()
    stop_pos = len(text)

    for stop_kw in stop_keywords:
        stop_match = re.search(rf'\n\s*{re.escape(stop_kw)}', text[start_pos:], re.IGNORECASE)
        if stop_match:
            stop_pos = start_pos + stop_match.start()
            break

    content = text[start_pos:stop_pos].strip()
    content = re.sub(r'\s*\d+\s*', ' ', content)
    content = re.sub(r'\s+', ' ', content)
    return content


def clean_text(text):
    """清理文本"""
    if not text:
        return ''
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


# ============================================================
# Excel 操作
# ============================================================
def get_oe_numbers(filepath):
    """从 A 列读取 OE 号"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    result = []
    for row in range(7, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            val = str(val).strip()
            if val:
                val = re.sub(r'[-\s]+[Pp]\d+$', '', val)
                result.append((row, val))
    return result


def write_results(filepath, output_path, results):
    """将解析结果写入 Excel"""
    shutil.copy2(filepath, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for row, data in results.items():
        if not data:
            continue
        for field, col in COL_MAP.items():
            val = data.get(field, '')
            if val:
                ws.cell(row=row, column=col, value=val)

    wb.save(output_path)


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("  离线测试 - 使用现有调试文件")
    print("=" * 60)

    # 定义测试数据：Row -> OE号 -> 调试文件名
    test_cases = [
        (7, 'MR249526', 'debug_reply_row7.txt'),
        (8, '22204-24010', 'debug_reply_row8.txt'),
        (9, '89543-28110', 'debug_reply_row9.txt'),
    ]

    results = {}
    ok = 0
    fail = 0

    for row, oe_number, filename in test_cases:
        print(f"\n{'='*60}")
        print(f"测试 Row {row}: {oe_number}")
        print(f"文件: {filename}")
        print(f"{'='*60}")

        # 读取调试文件
        filepath = os.path.join(DEBUG_DIR, filename)
        if not os.path.exists(filepath):
            print(f"    ✗ 文件不存在")
            fail += 1
            continue

        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        # 解析
        data = parse_deepseek_response(content, oe_number)
        if data:
            results[row] = data
            ok += 1
            print(f"    ✓ 解析成功!")
            print(f"      标题: {data['title'][:60]}...")
            print(f"      描述: {data['description'][:60]}...")
            print(f"      卖点1: {data['bullet1'][:60]}...")
        else:
            fail += 1
            print(f"    ✗ 解析失败")

    # 写入 Excel
    if results:
        try:
            write_results(EXCEL_PATH, OUTPUT_PATH, results)
            print(f"\n{'='*60}")
            print(f"✓ 结果已写入 Excel: {OUTPUT_PATH}")
            print(f"{'='*60}")
        except Exception as e:
            print(f"\n✗ 写入 Excel 失败: {e}")

    # 总结
    print(f"\n{'='*60}")
    print(f"  测试完成!")
    print(f"  成功: {ok} | 失败: {fail} | 总计: {len(test_cases)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
