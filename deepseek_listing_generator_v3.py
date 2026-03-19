# -*- coding: utf-8 -*-
"""
汽配零件 Listing 自动生成工具 (浏览器版 v3 - 优化版)
通过 Selenium 控制 Edge 浏览器操作 DeepSeek 网页版，生成亚马逊 listing 并回写 Excel。

修复内容：
1. 改进内容获取逻辑 - 只获取最新消息区域的文本
2. 优化解析器 - 支持多段落格式解析
3. 增强容错性 - 多种解析策略
4. 添加调试模式 - 保存原始响应便于排查
"""
import openpyxl
import shutil
import time
import re
import os
import sys

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ============================================================
# 配置区 - 根据需要修改
# ============================================================
EXCEL_PATH = r"C:\Users\Administrator\Desktop\test2.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\test2_output.xlsx"
DATA_START_ROW = 7
WAIT_TIMEOUT = 120          # 等待 DeepSeek 回复的最大秒数
REPLY_STABLE_COUNT = 4      # 连续N次文本不变则认为回复完成
CHECK_INTERVAL = 3          # 检查回复状态的间隔(秒)
ACTION_DELAY = 2            # 操作间隔秒数
MAX_RETRIES = 3             # 发送请求的最大重试次数
DEBUG_MODE = True           # 是否保存调试信息

# 列映射: 字段名 -> Excel 列号
COL_MAP = {
    'title': 7,              # G
    'description': 30,       # AD
    'bullet1': 31,           # AE
    'bullet2': 32,           # AF
    'bullet3': 33,           # AG
    'bullet4': 34,           # AH
    'bullet5': 35,           # AI
    'keyword1': 36,          # AJ
    'keyword2': 37,          # AK
    'keyword3': 38,          # AL
    'keyword4': 39,          # AM
    'keyword5': 40,          # AN
}


# ============================================================
# 改进的 DeepSeek 回复解析
# ============================================================
def parse_deepseek_response(text, oe_number):
    """
    解析 DeepSeek 的回复，支持多种格式。

    DeepSeek 输出格式分析：
    1. 表格行（tab分隔）: A G AD AE AF AG AH AI AJ AK AL AM AN
    2. 数据行（tab分隔）: OE号, Title, Description, Bullet1-5, Keyword1-5
    3. 然后是详细段落：
       - Vehicle Fitment: ... (描述)
       - Direct Fitment: ... (可能是描述的一部分或独立段落)
       - OEM Cross Reference: ... (通常是卖点2的一部分)
       - Precise Signal Output: ... (卖点3)
       - Plug & Play Installation: ... (卖点4)
       - Durable Construction: ... (卖点5的一部分)
       - 1-Year Warranty: ... (卖点5)
    4. 关键词: 最后的一行
    """
    if not text or len(text) < 100:
        return None

    result = {
        'oe': oe_number,
        'title': '',
        'description': '',
        'bullet1': '',
        'bullet2': '',
        'bullet3': '',
        'bullet4': '',
        'bullet5': '',
        'keyword1': '',
        'keyword2': '',
        'keyword3': '',
        'keyword4': '',
        'keyword5': '',
    }

    # ==================== 策略1: 解析 tab 分隔的表格行 ====================
    lines = text.split('\n')
    table_found = False

    for i, line in enumerate(lines):
        if '\t' in line:
            # 尝试提取表头
            parts = [p.strip().upper() for p in line.split('\t')]
            if len(parts) >= 13 and 'A' in parts and 'G' in parts and 'AD' in parts:
                # 找到表头，下一行应该是数据
                if i + 1 < len(lines):
                    data_line = lines[i + 1]
                    if '\t' in data_line:
                        fields = [f.strip() for f in data_line.split('\t')]
                        if len(fields) >= 13:
                            result['oe'] = fields[0]
                            result['title'] = fields[1]
                            result['description'] = fields[2]
                            result['bullet1'] = fields[3]
                            result['bullet2'] = fields[4]
                            result['bullet3'] = fields[5]
                            result['bullet4'] = fields[6]
                            result['bullet5'] = fields[7]
                            result['keyword1'] = fields[8]
                            result['keyword2'] = fields[9]
                            result['keyword3'] = fields[10]
                            result['keyword4'] = fields[11]
                            result['keyword5'] = fields[12]
                            table_found = True
                            break

    # 如果表格解析成功，返回
    if table_found and result['title']:
        print(f"    ✓ 通过表格格式解析成功")
        return result

    # ==================== 策略2: 解析段落格式 ====================
    print(f"    尝试段落格式解析...")

    # 提取标题（通常在 OE 号附近）
    # 格式1: [OE号] [Product Name] compatible with [Brand] [Model] [Year]
    title_patterns = [
        rf'{re.escape(oe_number)}\s+(.+?)\s+compatible\s+with',
        rf'{re.escape(oe_number)}\s+(.+?)(?:compatible|for|fits)',
    ]
    for pattern in title_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            # 尝试找完整的标题行
            title_line = match.group(0)
            # 扩展标题，包含 compatible with 后面的内容
            title_end = text.find(title_line) + len(title_line)
            # 寻找标题的结束位置（换行或句号）
            next_newline = text.find('\n', title_end)
            next_period = text.find('. ', title_end)
            if next_newline > 0 and (next_period == -1 or next_newline < next_period):
                title = text[:next_newline]
            elif next_period > 0:
                title = text[:next_period + 1]
            else:
                title = title_line[:200]  # 限制长度
            result['title'] = title.strip()
            break

    # 提取各个段落
    # Vehicle Fitment: ...
    vehicle_fitment_match = re.search(r'Vehicle\s+Fitment:\s*(.*?)(?=\n\s*(?:Direct\s+Fitment|OEM\s+Cross\s+Reference|Precise|Plug|Durable|1-Year|关键词|$))', text, re.IGNORECASE | re.DOTALL)
    if vehicle_fitment_match:
        result['description'] = clean_paragraph(vehicle_fitment_match.group(1))

    # Precise Signal Output: ... (卖点3)
    precise_match = re.search(r'Precise\s+Signal\s+Output:\s*(.*?)(?=\n\s*(?:Plug|Durable|1-Year|关键词|$))', text, re.IGNORECASE | re.DOTALL)
    if precise_match:
        result['bullet3'] = clean_paragraph(precise_match.group(1))
    else:
        # 尝试其他可能的模式
        for pattern in [r'Precise.*?:\s*(.*?)(?=\n\s*(?:Plug|Durable|1-Year))', r'性能.*?:\s*(.*?)(?=\n)']:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                result['bullet3'] = clean_paragraph(match.group(1))
                break

    # Plug & Play Installation: ... (卖点4)
    plug_match = re.search(r'Plug\s*&?\s*Play\s+Installation:\s*(.*?)(?=\n\s*(?:Durable|1-Year|关键词|$))', text, re.IGNORECASE | re.DOTALL)
    if plug_match:
        result['bullet4'] = clean_paragraph(plug_match.group(1))
    else:
        for pattern in [r'Plug.*?:\s*(.*?)(?=\n\s*(?:Durable|1-Year))', r'安装.*?:\s*(.*?)(?=\n)']:
            match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                result['bullet4'] = clean_paragraph(match.group(1))
                break

    # Durable Construction: ... (卖点5的一部分)
    durable_match = re.search(r'Durable\s+Construction:\s*(.*?)(?=\n\s*(?:1-Year|关键词|$))', text, re.IGNORECASE | re.DOTALL)
    if durable_match:
        durable_text = clean_paragraph(durable_match.group(1))
        # 合并到卖点5
        result['bullet5'] = durable_text

    # 1-Year Warranty: ... (卖点5的结尾)
    warranty_match = re.search(r'1-Year\s+Warranty:\s*(.*?)(?=\n\s*(?:关键词|写作思路|$))', text, re.IGNORECASE | re.DOTALL)
    if warranty_match:
        warranty_text = clean_paragraph(warranty_match.group(1))
        if result['bullet5']:
            result['bullet5'] = result['bullet5'] + ' ' + warranty_text
        else:
            result['bullet5'] = warranty_text

    # OEM Cross Reference: ... (卖点2的一部分)
    oem_match = re.search(r'OEM\s+Cross\s+Reference:\s*(.*?)(?=\n\s*(?:Precise|Plug|Durable|1-Year))', text, re.IGNORECASE | re.DOTALL)
    if oem_match:
        oem_text = clean_paragraph(oem_match.group(1))
        if not result['bullet2']:
            result['bullet2'] = oem_text
        else:
            result['bullet2'] = result['bullet2'] + ' ' + oem_text

    # 提取适配车型 (卖点1) - 从描述中提取第一句
    if result['description']:
        first_sentence = result['description'].split('.')[0]
        if 'compatible with' in first_sentence.lower() or 'compatible' in first_sentence.lower():
            result['bullet1'] = first_sentence.strip()
            # 从描述中移除
            result['description'] = result['description'][len(first_sentence):].strip()

    # 提取关键词
    # 通常在最后，格式：OE号 关键词1 关键词2 ...
    keywords_pattern = rf'{re.escape(oe_number)}\s+([\w\s\-\(\)]+?)\s*$'
    keywords_match = re.search(keywords_pattern, text, re.MULTILINE)
    if keywords_match:
        keywords_text = keywords_match.group(1).strip()
        keyword_list = [k.strip() for k in keywords_text.split() if k.strip()]
        for i, kw in enumerate(keyword_list[:5]):
            result[f'keyword{i+1}'] = kw

    # 验证是否提取到足够内容
    if result['title'] and (result['description'] or result['bullet1']):
        print(f"    ✓ 通过段落格式解析成功")
        return result
    else:
        print(f"    ✗ 段落格式解析失败")
        return None


def clean_paragraph(text):
    """清理段落文本"""
    if not text:
        return ''
    # 移除多余的空白和换行
    text = re.sub(r'\s+', ' ', text)
    # 移除引用标记
    text = re.sub(r'^\d+\s*', '', text)
    # 移除句尾的数字引用
    text = re.sub(r'\s*\d+\s*$', '', text)
    return text.strip()


# ============================================================
# Excel 操作
# ============================================================
def get_oe_numbers(filepath):
    """从 A 列读取 OE 号"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    result = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            val = str(val).strip()
            if val:
                # 去除数量后缀 P1/P4/P6/P8
                val = re.sub(r'[-\s]+[Pp]\d+$', '', val)
                result.append((row, val))
    return result


def write_results(filepath, output_path, results):
    """将解析结果写入 Excel"""
    shutil.copy2(filepath, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for row, data in results.items():
        if not data:
            continue
        for field, col in COL_MAP.items():
            val = data.get(field, '')
            if val:
                ws.cell(row=row, column=col, value=val)

    wb.save(output_path)


# ============================================================
# 浏览器操作 - 改进版本
# ============================================================
class DeepSeekBot:
    def __init__(self):
        self.driver = None

    def start(self):
        """启动 Edge"""
        print("启动 Edge 浏览器...")
        opts = Options()
        opts.add_argument('--start-maximized')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('detach', True)
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])

        self.driver = webdriver.Edge(options=opts)
        self.driver.get("https://chat.deepseek.com")
        print(f"已打开 DeepSeek，等待页面加载...")
        time.sleep(8)

    def _find_input(self):
        """查找输入框 - 改进版"""
        selectors = [
            # DeepSeek 可能的输入框选择器
            'textarea.dsb-textarea',
            'textarea[placeholder*="输入"]',
            'textarea[placeholder*="message"]',
            'div[contenteditable="true"]',
            'textarea',
            '#chat-input textarea',
            'div[class*="input"] textarea',
        ]
        for sel in selectors:
            try:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        return el
            except (NoSuchElementException, Exception):
                continue
        return None

    def _find_send_button(self):
        """查找发送按钮 - 改进版"""
        selectors = [
            'button[data-testid="send-button"]',
            'button[type="submit"]',
            'button[class*="send"]',
            'div[class*="send"] button',
            'button svg[class*="send"]',
            'button[aria-label*="send"]',
            'button[aria-label*="发送"]',
        ]
        for sel in selectors:
            try:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        return el
            except (NoSuchElementException, Exception):
                continue
        return None

    def _click_new_chat(self):
        """尝试开新对话"""
        selectors = [
            'button[class*="new"]',
            'a[class*="new"]',
            '[class*="new-chat"]',
            'div[class*="sidebar"] button:first-child',
            'button[aria-label*="new"]',
        ]
        for sel in selectors:
            try:
                btn = self.driver.find_element(By.CSS_SELECTOR, sel)
                if btn.is_displayed():
                    btn.click()
                    time.sleep(2)
                    return True
            except (NoSuchElementException, Exception):
                continue
        return False

    def _get_last_message_text(self):
        """
        获取最后一条消息的文本内容 - 关键改进！

        不再获取整个页面，而是只获取最后一条用户消息之后的内容
        """
        try:
            # 方法1: 尝试找到消息容器
            messages = self.driver.find_elements(By.CSS_SELECTOR, 'div[class*="message"], div[class*="chat-bubble"], div[class*="conversation"]')

            if messages:
                # 获取最后一条消息
                last_msg = messages[-1]
                # 只返回最后一条消息的文本
                return last_msg.text

            # 方法2: 尝试根据 OE 号定位
            body_text = self.driver.find_element(By.TAG_NAME, 'body').text

            # 找到最后一个 OE 号后面的内容
            # 假设 OE 号是我们发送的，最后一个 OE 号之后的就是回复
            oe_pattern = r'[A-Z0-9\-\.]+(?=[^\n]*?compatible|已收到您的OE号)'
            matches = list(re.finditer(oe_pattern, body_text))

            if matches:
                # 取最后一个匹配位置之后的内容
                last_pos = matches[-1].end()
                return body_text[last_pos:]

            return body_text

        except Exception as e:
            print(f"    获取消息失败: {str(e)}")
            return ""

    def send_and_get_reply(self, oe_number, max_retries=MAX_RETRIES):
        """发送 OE 号并获取完整回复 - 改进版"""
        retries = 0
        while retries <= max_retries:
            try:
                # 开新对话
                if not self._click_new_chat():
                    print("    警告: 无法开新对话，继续使用当前对话")

                # 找输入框
                input_el = self._find_input()
                if not input_el:
                    print(f"    找不到输入框 (尝试 {retries+1}/{max_retries})!")
                    retries += 1
                    time.sleep(2)
                    continue

                input_el.clear()
                input_el.click()
                time.sleep(0.5)

                # 输入 OE 号
                input_el.send_keys(oe_number)
                time.sleep(ACTION_DELAY)

                # 发送
                send_btn = self._find_send_button()
                if send_btn:
                    send_btn.click()
                else:
                    input_el.send_keys(Keys.RETURN)

                # 等待回复
                return self._wait_for_reply()

            except Exception as e:
                print(f"    发送失败 (尝试 {retries+1}/{max_retries}): {str(e)}")
                retries += 1
                time.sleep(2)
                if retries > max_retries:
                    print("    超过最大重试次数，跳过此条目")
                    return None

    def _wait_for_reply(self):
        """等待 DeepSeek 回复完成 - 改进版"""
        print(f"    等待回复 (最长 {WAIT_TIMEOUT}s)...")
        start = time.time()
        last_text = ""
        last_length = 0
        stable_count = 0

        while time.time() - start < WAIT_TIMEOUT:
            time.sleep(CHECK_INTERVAL)

            # 只获取最后一条消息
            current_text = self._get_last_message_text()
            current_length = len(current_text)

            # 检查是否有实质性内容
            if 'compatible' in current_text.lower() or 'fitment' in current_text.lower() or len(current_text) > 200:
                # 文本长度不再变化
                if current_length == last_length and current_length > 0:
                    stable_count += 1
                    if stable_count >= REPLY_STABLE_COUNT:
                        elapsed = int(time.time() - start)
                        print(f"    回复完成 ({elapsed}s, {current_length} 字符)")
                        return current_text
                else:
                    stable_count = 0
                    last_length = current_length

                # 显示进度
                if current_length > 0:
                    dots = "." * (int(time.time() - start) % 4)
                    print(f"    生成中{dots} ({int(time.time() - start)}s, {current_length} 字符)")

            last_text = current_text

        print(f"    超时，使用已收到的内容")
        return last_text

    def close(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("  汽配 Listing 自动生成工具 v3 (优化版)")
    print("  DeepSeek 浏览器自动化版")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"\nExcel 文件不存在: {EXCEL_PATH}")
        return

    # 读取 OE 号
    print(f"\n读取 Excel: {EXCEL_PATH}")
    oe_list = get_oe_numbers(EXCEL_PATH)
    print(f"找到 {len(oe_list)} 个 OE 号")
    for i, (row, oe) in enumerate(oe_list[:5]):
        print(f"  {i+1}. [Row {row}] {oe}")
    if len(oe_list) > 5:
        print(f"  ... 还有 {len(oe_list) - 5} 个")

    # 选择范围
    print(f"\n处理范围:")
    print(f"  1. 全部 ({len(oe_list)} 个)")
    print(f"  2. 测试前 3 个")
    print(f"  3. 自定义 (如 5-10)")
    choice = input("选择 (1/2/3): ").strip()

    if choice == '2':
        work = oe_list[:3]
    elif choice == '3':
        try:
            rng = input("范围 (如 5-10): ").strip().split('-')
            work = oe_list[int(rng[0])-1:int(rng[1])]
        except Exception:
            print("格式错误，处理全部")
            work = oe_list
    else:
        work = oe_list

    print(f"\n将处理 {len(work)} 个")

    # 启动浏览器
    bot = DeepSeekBot()
    try:
        bot.start()

        input("\n>>> 浏览器已打开，请确认 DeepSeek 已登录且 Prompt 已配置好。按 Enter 继续...")

        results = {}
        ok = 0
        fail = 0

        for i, (row, oe) in enumerate(work):
            print(f"\n[{i+1}/{len(work)}] OE: {oe} (Row {row})")

            reply = bot.send_and_get_reply(oe)
            if not reply:
                print("    未收到回复")
                fail += 1
                continue

            # 调试: 保存原始回复
            if DEBUG_MODE:
                debug_file = f"C:\\Users\\Administrator\\Desktop\\debug_v3_row{row}.txt"
                with open(debug_file, 'w', encoding='utf-8') as f:
                    f.write(f"OE Number: {oe}\n")
                    f.write(f"Row: {row}\n")
                    f.write("=" * 80 + "\n")
                    f.write(reply)
                print(f"    原始回复已保存: {debug_file}")

            # 解析
            data = parse_deepseek_response(reply, oe)
            if data and data.get('title'):
                results[row] = data
                ok += 1
                print(f"    ✓ 成功!")
                print(f"      标题: {data['title'][:60]}...")
                if data.get('description'):
                    print(f"      描述: {data['description'][:60]}...")
            else:
                fail += 1
                print(f"    ✗ 解析失败!")

            if i < len(work) - 1:
                time.sleep(ACTION_DELAY)

        # 写入
        if results:
            write_results(EXCEL_PATH, OUTPUT_PATH, results)
            print(f"\n✓ 结果已保存: {OUTPUT_PATH}")
        else:
            print(f"\n✗ 没有成功解析的数据")

        # 总结
        print(f"\n{'='*60}")
        print(f"  完成! 成功: {ok} | 失败: {fail} | 总计: {len(work)}")
        print(f"{'='*60}")

    except KeyboardInterrupt:
        print("\n用户中断")
    except Exception as e:
        print(f"\n出错: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("浏览器保持打开，可手动检查。")


if __name__ == "__main__":
    main()
