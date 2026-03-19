# -*- coding: utf-8 -*-
"""
汽配 Listing 自动生成工具 v4.2（table td 提取 - 简化等待逻辑）

改动：
1. _wait_for_reply 简化：只看 body 文本长度稳定即可
2. 发送前记录当前 table 数量，发送后等待新 table 出现
3. extract_from_table 找比发送前多的那个 table
"""
import openpyxl
import shutil
import time
import re
import os
import json

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# ============================================================
# 配置区
# ============================================================
EXCEL_PATH = r"C:\Users\Administrator\Desktop\test2.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\test2_output.xlsx"
DATA_START_ROW = 7
WAIT_TIMEOUT = 120
REPLY_STABLE_COUNT = 4
CHECK_INTERVAL = 3
ACTION_DELAY = 2
MAX_RETRIES = 3
DEBUG_MODE = True

COL_MAP = {
    'A': 1, 'G': 7, 'AD': 30,
    'AE': 31, 'AF': 32, 'AG': 33, 'AH': 34, 'AI': 35,
    'AJ': 36, 'AK': 37, 'AL': 38, 'AM': 39, 'AN': 40,
}

EXPECTED_HEADERS = ['A', 'G', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']

CHECKPOINT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "checkpoint.json"
)

# ============================================================
# 从 <table> <td> 提取数据
# ============================================================
def load_checkpoint():
    if os.path.exists(CHECKPOINT_PATH):
        try:
            with open(CHECKPOINT_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return set(data.get("done", []))
        except Exception:
            pass
    return set()

def save_checkpoint(done_rows):
    tmp = CHECKPOINT_PATH + ".tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump({"done": sorted(done_rows),
                   "updated": time.strftime("%Y-%m-%d %H:%M:%S")},
                  f, ensure_ascii=False, indent=2)
    os.replace(tmp, CHECKPOINT_PATH)

def clear_checkpoint():
    if os.path.exists(CHECKPOINT_PATH):
        os.remove(CHECKPOINT_PATH)

def extract_from_table(driver, oe_number, table_index=-1):
    """
    从 DeepSeek 页面的 table td 提取数据。
    table_index=-1 表示取最后一个 table。
    """
    try:
        tables = driver.find_elements(By.TAG_NAME, 'table')
        if not tables:
            print("    页面无 <table>")
            return None

        if table_index < 0:
            table_index = len(tables) - 1

        if table_index >= len(tables):
            print(f"    table_index={table_index} 超出范围({len(tables)})")
            return None

        table = tables[table_index]

        # 验证是目标 table（表头包含 A, G, AD）
        ths = table.find_elements(By.TAG_NAME, 'th')
        if not ths:
            first_tr = table.find_elements(By.TAG_NAME, 'tr')
            if first_tr:
                ths = first_tr[0].find_elements(By.TAG_NAME, 'td')

        header_texts = [t.text.strip().upper() for t in ths]
        if 'A' not in header_texts or 'G' not in header_texts or 'AD' not in header_texts:
            print(f"    table[{table_index}] 表头不匹配: {header_texts[:5]}...")
            return None

        # 提取数据行
        tbody = table.find_elements(By.TAG_NAME, 'tbody')
        data_trs = tbody[0].find_elements(By.TAG_NAME, 'tr') if tbody else []
        if not data_trs:
            # 没有 tbody，直接从 table 下的 tr 取
            all_trs = table.find_elements(By.TAG_NAME, 'tr')
            data_trs = all_trs[1:] if len(all_trs) > 1 else all_trs

        if not data_trs:
            print(f"    table[{table_index}] 无数据行")
            return None

        tds = data_trs[0].find_elements(By.TAG_NAME, 'td')
        if not tds or len(tds) < 10:
            print(f"    table[{table_index}] 只有 {len(tds)} 个 td")
            return None

        # 提取文本
        td_texts = []
        for td in tds:
            text = td.text.strip()
            text = re.sub(r'\s+', ' ', text)
            td_texts.append(text)

        num_cols = len(td_texts)
        print(f"    table[{table_index}] {num_cols} 列, 表头 {len(header_texts)} 个")

        result = {
            'A': '', 'G': '', 'AD': '', 'AE': '', 'AF': '', 'AG': '',
            'AH': '', 'AI': '', 'AJ': '', 'AK': '', 'AL': '', 'AM': '', 'AN': '',
        }

        if num_cols >= 14:
            # 14 列：td[8] 对应表头 AJ，直接映射，不合并
            map14 = {0:'A',1:'G',2:'AD',3:'AE',4:'AF',5:'AG',6:'AH',7:'AI',
                     8:'AJ',9:'AK',10:'AL',11:'AM',12:'AN'}
            for idx, col in map14.items():
                if idx < num_cols: result[col] = td_texts[idx]
            # td[13] 无对应表头，丢弃
        elif num_cols >= 13:
            for i, col in enumerate(EXPECTED_HEADERS):
                if i < num_cols: result[col] = td_texts[i]
        else:
            return None

        if result['G'] and len(result['G']) > 5:
            filled = sum(1 for v in result.values() if v)
            print(f"    OK! {filled}/13 字段")
            for col in ['G','AD','AE','AF','AG','AH','AI','AJ','AK']:
                v = result.get(col,'')
                if v: print(f"      {col}: {len(v):4}c")
            return result

        return None

    except Exception as e:
        print(f"    table 异常: {str(e)}")
        return None


# ============================================================
# 文本解析 fallback
# ============================================================
HEADER_RE = re.compile(r'^A\s+G\s+AD\s+AE\s+AF\s+AG\s+AH\s+AI\s+AJ\s+AK\s+AL\s+AM\s+AN\s*$', re.I)

def parse_deepseek_text(text, oe_number):
    if not text or len(text) < 100: return None
    lines = text.split('\n')
    # 找最后一个表头行的数据
    last_hdr = -1
    for i, line in enumerate(lines):
        if HEADER_RE.match(line.strip()):
            last_hdr = i
    if last_hdr < 0: return None
    data_lines = []
    for i in range(last_hdr+1, len(lines)):
        s = lines[i].strip()
        if HEADER_RE.match(s): break
        if re.match(r'^\d+$', s): continue
        if s: data_lines.append(s)
    merged = ' '.join(data_lines)
    merged = re.sub(r'\)\s+[\d\s]+([,;.])', r')\1', merged)
    merged = re.sub(r'\s+', ' ', merged).strip()
    if len(merged) < 50: return None

    result = {'A':oe_number,'G':'','AD':'','AE':'','AF':'','AG':'','AH':'','AI':'',
              'AJ':'','AK':'','AL':'','AM':'','AN':''}
    pos = merged.find(oe_number)
    if pos < 0: return None
    vf = re.search(r'Vehicle\s+Fitment:', merged, re.I)
    if not vf: return None
    result['G'] = merged[pos+len(oe_number):vf.start()].strip()
    result['G'] = re.sub(r'^\d{4,5}[-]\d{4,5}\s*','',result['G']).strip()
    result['G'] = re.sub(r'\s+\d{4,5}[-]\d{4,5}\s*$','',result['G']).strip()
    if not result['G']: return None
    oes = list(re.finditer(re.escape(oe_number), merged))
    last_end = oes[-1].end() if oes else len(merged)
    vf_end = vf.end()
    bps = [r'Verify\s+vehicle',r'Direct\s+fit\s+for',r'Replaces\s+OEM',
           r'Provides\s+accurate',r'Measures\s+intake',r'Plug\s*&?\s*play',
           r'Delivers\s+accurate',r'Durable',r'Corrosion-resistant',
           r'100%\s+tested',r'1-?year\s+warranty',r'Meets\s+OE',
           r'Meets\s+Toyota',r'No\s+splicing',r'Precise\s+[SA]']
    found = []
    for p in bps:
        for m in re.finditer(p, merged[vf_end:], re.I):
            found.append(vf_end + m.start())
    found = sorted(set(found))
    ff = []
    for f in found:
        if not ff or f - ff[-1] > 5: ff.append(f)
    ff = [f for f in ff if vf_end < f < last_end]
    if ff:
        result['AD'] = re.sub(r'\s*\.?\s*$','',merged[vf_end:ff[0]].strip())
    else:
        result['AD'] = re.sub(r'\s*\.?\s*$','',merged[vf_end:last_end].strip())
    for i in range(min(5, len(ff))):
        s = ff[i]
        e = ff[i+1] if i+1 < len(ff) else last_end
        col = ['AE','AF','AG','AH','AI'][i]
        result[col] = merged[s:e].strip()
    if last_end < len(merged):
        kw = merged[last_end:].strip()
        kw = re.sub(r'(已读取|Read).*$','',kw,flags=re.DOTALL).strip()
        kws = [k.strip() for k in kw.split() if len(k.strip())>1]
        for i,k in enumerate(kws[:5]):
            result[['AJ','AK','AL','AM','AN'][i]] = k
    return result if result['G'] else None


# ============================================================
# Excel
# ============================================================
def get_oe_numbers(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    r = []
    for row in range(DATA_START_ROW, ws.max_row+1):
        v = ws.cell(row=row, column=1).value
        if v:
            v = str(v).strip()
            v = re.sub(r'[-\s]+[Pp]\d+$','',v)
            if v: r.append((row, v))
    return r

def write_single_row(fp, out, row, data):
    if not os.path.exists(out):
        shutil.copy2(fp, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    for cn, num in COL_MAP.items():
        v = data.get(cn, '')
        if v:
            ws.cell(row=row, column=num, value=v)
    wb.save(out)


# ============================================================
# 浏览器
# ============================================================
class DeepSeekBot:
    def __init__(self):
        self.driver = None

    def start(self):
        print("启动 Edge...")
        opts = Options()
        opts.add_argument('--start-maximized')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('detach', True)
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])
        self.driver = webdriver.Edge(options=opts)
        self.driver.get("https://chat.deepseek.com")
        print("等待加载...")
        time.sleep(8)

    def _find_input(self):
        for sel in ['textarea.dsb-textarea','textarea[placeholder*="输入"]',
                     'textarea[placeholder*="message"]','div[contenteditable="true"]',
                     'textarea','div[class*="input"] textarea']:
            try:
                for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.is_enabled(): return el
            except: continue
        return None

    def _find_send(self):
        for sel in ['button[data-testid="send-button"]','button[type="submit"]',
                     'button[class*="send"]','div[class*="send"] button',
                     'button[aria-label*="send"]','button[aria-label*="发送"]']:
            try:
                for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.is_enabled(): return el
            except: continue
        return None

    def _click_new_chat(self):
        for sel in ['button[class*="new"]','a[class*="new"]','[class*="new-chat"]',
                     'div[class*="sidebar"] button:first-child','button[aria-label*="new"]']:
            try:
                btn = self.driver.find_element(By.CSS_SELECTOR, sel)
                if btn.is_displayed():
                    btn.click()
                    time.sleep(2)
                    return True
            except: continue
        return False

    def _count_tables(self):
        """统计页面上匹配的 table 数量"""
        try:
            tables = self.driver.find_elements(By.TAG_NAME, 'table')
            count = 0
            for t in tables:
                ths = t.find_elements(By.TAG_NAME, 'th')
                ht = [th.text.strip().upper() for th in ths]
                if 'A' in ht and 'G' in ht and 'AD' in ht:
                    count += 1
            return count
        except:
            return 0

    def send_and_wait(self, oe_number):
        """发送 OE 号并等待回复（返回 (True, table_index) 或 (False, -1)）"""
        for attempt in range(MAX_RETRIES + 1):
            # 开新对话
            self._click_new_chat()

            # 记录当前 table 数量
            tables_before = self._count_tables()

            # 发送
            input_el = self._find_input()
            if not input_el:
                print(f"    找不到输入框 ({attempt+1})")
                time.sleep(3)
                continue

            input_el.clear()
            input_el.click()
            time.sleep(0.5)
            input_el.send_keys(oe_number)
            time.sleep(ACTION_DELAY)

            send_btn = self._find_send()
            if send_btn:
                send_btn.click()
            else:
                input_el.send_keys(Keys.RETURN)

            # 等待回复
            print(f"    等待回复 (当前 {tables_before} 个 table)...")
            start = time.time()
            last_len = -1
            stable = 0

            while time.time() - start < WAIT_TIMEOUT:
                time.sleep(CHECK_INTERVAL)

                try:
                    body_len = len(self.driver.find_element(By.TAG_NAME, 'body').text)
                    tables_now = self._count_tables()
                except:
                    body_len = 0
                    tables_now = 0

                marker = tables_now * 100000 + body_len
                if marker == last_len and body_len > 0:
                    stable += 1
                    if stable >= REPLY_STABLE_COUNT:
                        elapsed = int(time.time() - start)
                        print(f"    回复完成 ({elapsed}s, {tables_now} table, {body_len}c)")

                        # 新增了 table？
                        if tables_now > tables_before:
                            new_table_idx = -1
                            all_tables = self.driver.find_elements(By.TAG_NAME, 'table')
                            # 从后往前找新增的那个
                            for i in range(len(all_tables)-1, -1, -1):
                                t = all_tables[i]
                                ths = t.find_elements(By.TAG_NAME, 'th')
                                ht = [th.text.strip().upper() for th in ths]
                                if 'A' in ht and 'G' in ht:
                                    if i >= tables_before - 1:  # 允许 ±1 的误差
                                        new_table_idx = i
                                    break
                            return (True, new_table_idx)
                        else:
                            # 没有 table 变化，用最后一个
                            all_tables = self.driver.find_elements(By.TAG_NAME, 'table')
                            return (True, len(all_tables) - 1)
                else:
                    stable = 0
                last_len = marker

                dots = "." * (int(time.time() - start) % 4)
                print(f"    生成中{dots} ({int(time.time()-start)}s, t={tables_now})")

            print(f"    超时")

        return (False, -1)

    def close(self):
        if self.driver:
            try: self.driver.quit()
            except: pass


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("  汽配 Listing 自动生成工具 v4.2")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"Excel 不存在: {EXCEL_PATH}"); return

    print(f"\n读取 Excel: {EXCEL_PATH}")
    oe_list = get_oe_numbers(EXCEL_PATH)
    print(f"找到 {len(oe_list)} 个 OE 号")

    print(f"\n  1. 全部 ({len(oe_list)})")
    print(f"  2. 测试前 3 个")
    choice = input("选择 (1/2): ").strip()
    work = oe_list[:3] if choice == '2' else oe_list
    print(f"\n处理 {len(work)} 个")

    bot = DeepSeekBot()
    try:
        done_rows = load_checkpoint()
        if done_rows:
        print(f"✅ 发现断点记录，已完成 {len(done_rows)} 行，将自动跳过")
        pending = [(row, oe) for row, oe in oe_list if row not in done_rows]
        bot.start()
        input("\n>>> 确认 DeepSeek 已登录且 Prompt 配置好。按 Enter 继续...")

        results = {}
        ok = fail = 0
        stats = {'table':0, 'text':0}

        for i, (row, oe) in enumerate(work):
            print(f"\n{'='*60}")
            print(f"[{i+1}/{len(work)}] OE: {oe} (Row {row})")
            print(f"{'='*60}")

            reply_ok, table_idx = bot.send_and_wait(oe)
            if not reply_ok:
                print("    X 未收到回复"); fail += 1; continue

            # 调试：保存 table HTML
            if DEBUG_MODE:
                try:
                    all_t = bot.driver.find_elements(By.TAG_NAME, 'table')
                    if all_t:
                        idx = table_idx if 0 <= table_idx < len(all_t) else len(all_t)-1
                        html = all_t[idx].get_attribute('outerHTML')
                        df = os.path.join(r"C:\Users\Administrator\Desktop\deepseek_listing_tool", f"debug_v42_row{row}.html")
                        with open(df, 'w', encoding='utf-8') as f:
                            f.write(f"OE:{oe} Row:{row} Tables:{len(all_t)} TableIdx:{idx}\n{'='*80}\n")
                            f.write(html)
                except: pass

            # 提取数据
            data = extract_from_table(bot.driver, oe, table_idx)
            if data:
                write_single_row(EXCEL_PATH, OUTPUT_PATH, row, data)
                done_rows.add(row)
                save_checkpoint(done_rows)
                ok += 1
                print(f"  ✅ 已写入 Row {row} ({len(done_rows)}/{len(oe_list)})")
            else:
                print("    -> fallback 文本解析")
                try:
                    bt = bot.driver.find_element(By.TAG_NAME, 'body').text
                    data = parse_deepseek_text(bt, oe)
                    if data: stats['text'] += 1
                except Exception as e:
                    print(f"    fallback 失败: {e}")

            if data:
                results[row] = data; ok += 1
                print(f"    OK!")
                for c in ['G','AD','AE','AF','AG','AH','AI','AJ','AK']:
                    v = data.get(c,'')
                    if v: print(f"      {c}: {len(v):4}c | {v[:65]}...")
            else:
                fail += 1
                print(f"    X 失败!")

            if i < len(work)-1: time.sleep(ACTION_DELAY)

        if results:
            write_results(EXCEL_PATH, OUTPUT_PATH, results)

        print(f"\n{'='*60}")
        print(f"  完成! 成功:{ok} 失败:{fail} 总计:{len(work)}")
        print(f"  table={stats['table']} text={stats['text']}")
        print(f"{'='*60}")

    except KeyboardInterrupt:
        print("\n中断")
    except Exception as e:
        print(f"\n出错: {e}")
        import traceback; traceback.print_exc()
    finally:
        print("\n浏览器保持打开。")


if __name__ == "__main__":
    main()