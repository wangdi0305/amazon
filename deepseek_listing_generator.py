# -*- coding: utf-8 -*-
"""
汽配零件 Listing 自动生成工具 (浏览器版 v2)
通过 Selenium 控制 Edge 浏览器操作 DeepSeek 网页版，生成亚马逊 listing 并回写 Excel。

修复: 解析逻辑匹配 DeepSeek 实际的 tab 分隔表格输出格式
"""
import openpyxl
import shutil
import time
import re
import os
import sys

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ============================================================
# 配置区 - 根据需要修改
# ============================================================
EXCEL_PATH = r"C:\Users\Administrator\Desktop\test2.xlsx"
OUTPUT_PATH = r"C:\Users\Administrator\Desktop\test2_output.xlsx"
DATA_START_ROW = 7
WAIT_TIMEOUT = 120          # 等待 DeepSeek 回复的最大秒数
REPLY_STABLE_COUNT = 4      # 连续N次文本不变则认为回复完成
CHECK_INTERVAL = 3          # 检查回复状态的间隔(秒)
ACTION_DELAY = 2            # 操作间隔秒数
MAX_RETRIES = 3             # 发送请求的最大重试次数

# 列映射: 字段名 -> Excel 列号
COL_MAP = {
    'title': 7,              # G
    'description': 30,       # AD
    'bullet1': 31,           # AE
    'bullet2': 32,           # AF
    'bullet3': 33,           # AG
    'bullet4': 34,           # AH
    'bullet5': 35,           # AI
    'keyword1': 36,          # AJ
    'keyword2': 37,          # AK
    'keyword3': 38,          # AL
    'keyword4': 39,          # AM
    'keyword5': 40,          # AN
}


# ============================================================
# 解析 DeepSeek 回复 (tab 分隔表格格式)
# ============================================================
def parse_deepseek_table(text):
    """
    解析 DeepSeek 的 tab 分隔表格输出。

    DeepSeek 输出结构:
      1. 描述段落
      2. 空行
      3. 表头: A\\tG\\tAD\\tAE\\tAF\\tAG\\tAH\\tAI\\tAJ\\tAK\\tAL\\tAM\\tAN
      4. 数据行: OE\\tTitle\\tDesc\\tB1\\tB2\\tB3\\tB4\\tB5\\tK1\\tK2\\tK3\\tK4\\tK5
      5. 空行
      6. 写作思路说明段落
    """
    if not text or len(text) < 50:
        return None

    lines = text.split('\n')

    # 方案1: 查找表头行
    for i, line in enumerate(lines):
        # 表头包含 A, G, AD, AE 等列标识，用 tab 分隔
        if '\t' in line:
            parts = [p.strip().upper() for p in line.split('\t')]
            if len(parts) >= 12 and 'A' in parts and 'G' in parts and 'AD' in parts:
                # 找到表头，取下一行为数据
                for j in range(i + 1, min(i + 3, len(lines))):
                    data_line = lines[j].strip()
                    if data_line and '\t' in data_line:
                        fields = data_line.split('\t')
                        if len(fields) >= 12:
                            return _map_fields(fields)

    # 方案2: 查找第一个含大量 tab 且以 OE 号开头的行
    for line in lines:
        if '\t' in line:
            fields = line.split('\t')
            if len(fields) >= 12:
                first = fields[0].strip()
                # OE 号格式: 字母数字+连字符
                if re.match(r'^[A-Z0-9][A-Z0-9\-\.]+$', first, re.IGNORECASE):
                    return _map_fields(fields)

    return None


def _map_fields(fields):
    """将 tab 分割的字段映射到标准 key"""
    def get(idx):
        return fields[idx].strip() if idx < len(fields) else ''

    return {
        'oe': get(0),
        'title': get(1),
        'description': get(2),
        'bullet1': get(3),
        'bullet2': get(4),
        'bullet3': get(5),
        'bullet4': get(6),
        'bullet5': get(7),
        'keyword1': get(8),
        'keyword2': get(9),
        'keyword3': get(10),
        'keyword4': get(11),
        'keyword5': get(12),
    }


# ============================================================
# Excel 操作
# ============================================================
def get_oe_numbers(filepath):
    """从 A 列读取 OE 号"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    result = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            val = str(val).strip()
            if val:
                # 去除数量后缀 P1/P4/P6/P8
                val = re.sub(r'[-\s]+[Pp]\d+$', '', val)
                result.append((row, val))
    return result


def write_results(filepath, output_path, results):
    """将解析结果写入 Excel"""
    shutil.copy2(filepath, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for row, data in results.items():
        if not data:
            continue
        for field, col in COL_MAP.items():
            val = data.get(field, '')
            if val:
                ws.cell(row=row, column=col, value=val)

    wb.save(output_path)


# ============================================================
# 浏览器操作
# ============================================================
class DeepSeekBot:
    def __init__(self):
        self.driver = None

    def start(self):
        """启动 Edge"""
        print("启动 Edge 浏览器...")
        opts = Options()
        opts.add_argument('--start-maximized')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('detach', True)
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])

        self.driver = webdriver.Edge(options=opts)
        self.driver.get("https://chat.deepseek.com")
        print(f"已打开 DeepSeek，等待页面加载...")
        time.sleep(8)

    def _find_input(self):
        """查找输入框"""
        for sel in [
            'textarea.dsb-textarea',
            'textarea',
            'div[contenteditable="true"]',
            '#chat-input textarea',
        ]:
            try:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed():
                        return el
            except NoSuchElementException:
                continue
        return None

    def _find_send_button(self):
        """查找发送按钮"""
        for sel in [
            'button[data-testid="send-button"]',
            'div[class*="send"] button',
            'button[class*="send"]',
            'div[class*="toolbar"] button:last-child',
        ]:
            try:
                el = self.driver.find_element(By.CSS_SELECTOR, sel)
                if el.is_displayed():
                    return el
            except NoSuchElementException:
                continue
        return None

    def _click_new_chat(self):
        """尝试开新对话"""
        for sel in [
            'button[class*="new"]',
            'a[class*="new"]',
            '[class*="new-chat"]',
            'div[class*="sidebar"] button:first-child',
        ]:
            try:
                btn = self.driver.find_element(By.CSS_SELECTOR, sel)
                if btn.is_displayed():
                    btn.click()
                    time.sleep(2)
                    return True
            except (NoSuchElementException, Exception):
                continue
        return False

    def send_and_get_reply(self, oe_number, max_retries=MAX_RETRIES):
        """发送 OE 号并获取完整回复"""
        retries = 0
        while retries <= max_retries:
            try:
                # 开新对话
                self._click_new_chat()

                # 找输入框
                input_el = self._find_input()
                if not input_el:
                    print(f"    找不到输入框 (尝试 {retries+1}/{max_retries})!")
                    retries += 1
                    time.sleep(2)
                    continue

                input_el.clear()
                input_el.click()
                time.sleep(0.5)

                # 输入 OE 号
                input_el.send_keys(oe_number)
                time.sleep(ACTION_DELAY)

                # 发送
                send_btn = self._find_send_button()
                if send_btn:
                    send_btn.click()
                else:
                    input_el.send_keys(Keys.RETURN)

                # 等待回复
                return self._wait_for_reply()

            except Exception as e:
                print(f"    发送失败 (尝试 {retries+1}/{max_retries}): {str(e)}")
                retries += 1
                time.sleep(2)
                if retries > max_retries:
                    print("    超过最大重试次数，跳过此条目")
                    return None

    def _get_all_text(self):
        """获取页面所有文本内容"""
        try:
            return self.driver.find_element(By.TAG_NAME, 'body').text
        except Exception:
            return ""

    def _wait_for_reply(self):
        """等待 DeepSeek 回复完成"""
        print(f"    等待回复 (最长 {WAIT_TIMEOUT}s)...")
        start = time.time()
        last_text = ""
        stable = 0

        while time.time() - start < WAIT_TIMEOUT:
            time.sleep(CHECK_INTERVAL)
            full_text = self._get_all_text()

            if full_text == last_text and len(full_text) > 100:
                stable += 1
                if stable >= REPLY_STABLE_COUNT:
                    elapsed = int(time.time() - start)
                    print(f"    回复完成 ({elapsed}s)")
                    return full_text
            else:
                stable = 0
                last_text = full_text
                if full_text:
                    dots = "." * (int(time.time() - start) % 4)
                    print(f"    生成中{dots} ({int(time.time() - start)}s)")

        print(f"    超时，使用已收到的内容")
        return last_text

    def close(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("  汽配 Listing 自动生成工具 v2")
    print("  DeepSeek 浏览器自动化版")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"\nExcel 文件不存在: {EXCEL_PATH}")
        return

    # 读取 OE 号
    print(f"\n读取 Excel: {EXCEL_PATH}")
    oe_list = get_oe_numbers(EXCEL_PATH)
    print(f"找到 {len(oe_list)} 个 OE 号")
    for i, (row, oe) in enumerate(oe_list[:5]):
        print(f"  {i+1}. [Row {row}] {oe}")
    if len(oe_list) > 5:
        print(f"  ... 还有 {len(oe_list) - 5} 个")

    # 选择范围
    print(f"\n处理范围:")
    print(f"  1. 全部 ({len(oe_list)} 个)")
    print(f"  2. 测试前 3 个")
    print(f"  3. 自定义 (如 5-10)")
    choice = input("选择 (1/2/3): ").strip()

    if choice == '2':
        work = oe_list[:3]
    elif choice == '3':
        try:
            rng = input("范围 (如 5-10): ").strip().split('-')
            work = oe_list[int(rng[0])-1:int(rng[1])]
        except Exception:
            print("格式错误，处理全部")
            work = oe_list
    else:
        work = oe_list

    print(f"\n将处理 {len(work)} 个")

    # 启动浏览器
    bot = DeepSeekBot()
    try:
        bot.start()

        input("\n>>> 浏览器已打开，请确认 DeepSeek 已登录且 Prompt 已配置好。按 Enter 继续...")

        results = {}
        ok = 0
        fail = 0

        for i, (row, oe) in enumerate(work):
            print(f"\n[{i+1}/{len(work)}] OE: {oe} (Row {row})")

            reply = bot.send_and_get_reply(oe)
            if not reply:
                print("    未收到回复")
                fail += 1
                continue

            # 解析
            data = parse_deepseek_table(reply)
            if data and data.get('title'):
                results[row] = data
                ok += 1
                print(f"    成功! Title: {data['title'][:70]}")
            else:
                fail += 1
                # 保存原始回复供调试
                debug_file = f"C:\\Users\\Administrator\\Desktop\\debug_reply_row{row}.txt"
                with open(debug_file, 'w', encoding='utf-8') as f:
                    f.write(reply)
                print(f"    解析失败! 原始回复已保存到: {debug_file}")

            if i < len(work) - 1:
                time.sleep(ACTION_DELAY)

        # 写入
        if results:
            write_results(EXCEL_PATH, OUTPUT_PATH, results)
            print(f"\n结果已保存: {OUTPUT_PATH}")

        # 总结
        print(f"\n{'='*60}")
        print(f"  完成! 成功: {ok} | 失败: {fail} | 总计: {len(work)}")
        print(f"{'='*60}")

    except KeyboardInterrupt:
        print("\n用户中断")
    except Exception as e:
        print(f"\n出错: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("浏览器保持打开，可手动检查。")


if __name__ == "__main__":
    main()
